import os
import json
import uuid
import logging
import time
import threading
import queue
from io import BytesIO
from base64 import b64encode
from datetime import datetime
from collections import defaultdict

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')
logger = logging.getLogger(__name__)

from flask import Flask, render_template, request, session, redirect, url_for, jsonify, send_file, Response
from flask_socketio import SocketIO, emit, join_room
from sqlalchemy import create_engine, Column, String, Integer, DateTime, Boolean, JSON, Float, text
from sqlalchemy.orm import declarative_base, sessionmaker
import qrcode

app = Flask(__name__, template_folder='.')
app.secret_key = os.environ.get('SECRET_KEY', 'cs671-viva-g27')
app.config['SESSION_COOKIE_DOMAIN'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading',
                    ping_timeout=60, ping_interval=25, max_http_buffer_size=1_000_000)

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'samiscrazy')

DB_PATH = os.environ.get('DB_PATH', 'quiz_platform.db')
engine = create_engine(
    f'sqlite:///{DB_PATH}',
    connect_args={"check_same_thread": False, "timeout": 30},
    pool_size=10, max_overflow=20, pool_timeout=30, pool_pre_ping=True
)
with engine.connect() as conn:
    conn.execute(text("PRAGMA journal_mode=WAL"))
    conn.execute(text("PRAGMA synchronous=NORMAL"))
    conn.execute(text("PRAGMA cache_size=10000"))
    conn.execute(text("PRAGMA temp_store=MEMORY"))
    conn.commit()

Base = declarative_base()
SessionLocal = sessionmaker(bind=engine)

class Quiz(Base):
    __tablename__ = 'quizzes'
    id = Column(String(36), primary_key=True)
    title = Column(String(255), nullable=False)
    pin = Column(String(6), unique=True, nullable=False)
    admin_id = Column(String(36), nullable=False)
    questions_data = Column(JSON, nullable=False)
    active_question_id = Column(Integer, nullable=True)
    active_question_open_time = Column(DateTime, nullable=True)
    state = Column(String(50), default='idle')
    created_at = Column(DateTime, default=datetime.utcnow)

class QuizSession(Base):
    __tablename__ = 'quiz_sessions'
    id = Column(String(36), primary_key=True)
    quiz_id = Column(String(36), nullable=False)
    username = Column(String(255), nullable=False)
    correct_count = Column(Integer, default=0)
    speed_points = Column(Integer, default=0)
    strikes = Column(Integer, default=0)
    attended = Column(Boolean, default=True)
    started_at = Column(DateTime, default=datetime.utcnow)
    completed_at = Column(DateTime, nullable=True)

class Answer(Base):
    __tablename__ = 'answers'
    id = Column(String(36), primary_key=True)
    quiz_id = Column(String(36), nullable=False)
    session_id = Column(String(36), nullable=False)
    question_id = Column(Integer, nullable=False)
    user_answer = Column(String, nullable=True)
    is_correct = Column(Boolean, nullable=False)
    partial_score = Column(Integer, default=0)
    speed_points = Column(Integer, default=0)
    time_taken = Column(Float, nullable=True)
    strike_on_question = Column(Boolean, default=False)
    answered_at = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(engine)

# ─── IN-MEMORY LEADERBOARD CACHE ───────────────────────────────────────────────
# {quiz_id: {session_id: {username, correct, speed_points, strikes}}}
_lb_cache = defaultdict(dict)
_lb_lock = threading.Lock()
_lb_pending = set()
_lb_last_broadcast = {}
_LB_THROTTLE = 1.0  # max 1 broadcast/second per quiz

def _lb_update(quiz_id, sess_id, username, correct, speed, strikes):
    with _lb_lock:
        _lb_cache[quiz_id][sess_id] = {'username': username, 'correct': correct,
                                        'speed_points': speed, 'strikes': strikes}
        _lb_pending.add(quiz_id)

def _get_sorted_lb(quiz_id):
    with _lb_lock:
        entries = list(_lb_cache[quiz_id].values())
    entries.sort(key=lambda x: (-x['correct'], -x['speed_points']))
    return entries

def _lb_broadcaster():
    while True:
        time.sleep(0.25)
        now = time.time()
        with _lb_lock:
            to_send = set(_lb_pending)
            _lb_pending.clear()
        for quiz_id in to_send:
            if now - _lb_last_broadcast.get(quiz_id, 0) >= _LB_THROTTLE:
                _lb_last_broadcast[quiz_id] = now
                lb = _get_sorted_lb(quiz_id)
                socketio.emit('leaderboard_update', {'leaderboard': lb}, room=quiz_id)

threading.Thread(target=_lb_broadcaster, daemon=True).start()

def _prime_lb_cache(quiz_id):
    with _lb_lock:
        if quiz_id in _lb_cache:
            return
    db = SessionLocal()
    sessions = db.query(QuizSession).filter(QuizSession.quiz_id == quiz_id).all()
    db.close()
    with _lb_lock:
        for s in sessions:
            _lb_cache[quiz_id][s.id] = {'username': s.username, 'correct': s.correct_count,
                                          'speed_points': s.speed_points, 'strikes': s.strikes}

# ─── ASYNC WRITE QUEUE ─────────────────────────────────────────────────────────
_write_queue = queue.Queue()

def _write_worker():
    while True:
        fn = _write_queue.get()
        try:
            fn()
        except Exception as e:
            logger.error(f"Write error: {e}")
        finally:
            _write_queue.task_done()

threading.Thread(target=_write_worker, daemon=True).start()

# ─── HELPERS ───────────────────────────────────────────────────────────────────

def gen_pin():
    return str(uuid.uuid4())[:6].upper()

def gen_qr(data):
    qr = qrcode.QRCode(version=1, box_size=8, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return b64encode(buf.getvalue()).decode()

def calc_speed_points(time_taken, time_limit=None):
    max_time = time_limit if time_limit else 30
    if time_taken is None or time_taken <= 0:
        return 500
    return int(max(0, 1 - time_taken / max_time) * 1000)

def check_answer_detailed(q, user_answer):
    t = q.get('type', 'mcq')
    if t == 'mcq':
        c = user_answer == str(q.get('correct_option'))
        return c, (1 if c else 0), not c
    elif t == 'true_false':
        c = (user_answer.lower() == 'true') == q.get('correct_answer')
        return c, (1 if c else 0), not c
    elif t == 'fill_up':
        c = user_answer.lower().strip() == q.get('correct_answer', '').lower().strip()
        return c, (1 if c else 0), not c
    elif t == 'multi_correct':
        try:
            user_set = set(int(x) for x in user_answer.split(',') if x.strip())
            correct_set = set(q.get('correct_options', []))
            wrong_set = set(range(len(q.get('options', [])))) - correct_set
            score = max(0, len(user_set & correct_set) - len(user_set & wrong_set))
            any_wrong = bool(user_set & wrong_set)
            return user_set == correct_set, score, any_wrong
        except:
            return False, 0, True
    return False, 0, True

def _get_correct_answer_str(q):
    t = q.get('type', 'mcq')
    if t == 'mcq':
        idx = q.get('correct_option', 0)
        opts = q.get('options', [])
        return opts[idx] if idx < len(opts) else str(idx)
    elif t == 'true_false':
        return str(q.get('correct_answer', '')).lower()
    elif t == 'fill_up':
        return q.get('correct_answer', '')
    elif t == 'multi_correct':
        idxs = q.get('correct_options', [])
        opts = q.get('options', [])
        return ', '.join(opts[i] for i in idxs if i < len(opts))
    return ''

# ─── ROUTES - STUDENT ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('login.html', prefill_pin=request.args.get('pin', ''), error=None)

@app.route('/join', methods=['GET', 'POST'])
def join():
    if request.method == 'GET':
        pin = request.args.get('pin', '').strip().upper()
        return redirect(url_for('index', pin=pin) if pin else url_for('index'))

    quiz_pin = request.form.get('quiz_pin', '').strip().upper()
    username = request.form.get('username', '').strip()

    if not quiz_pin or not username:
        return render_template('login.html', error='PIN and username required.', prefill_pin=quiz_pin)

    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.pin == quiz_pin).first()
    if not quiz:
        db.close()
        return render_template('login.html', error='Invalid quiz PIN.', prefill_pin=quiz_pin)

    quiz_id = quiz.id
    existing = db.query(QuizSession).filter(
        QuizSession.quiz_id == quiz_id,
        QuizSession.username.ilike(username)
    ).order_by(QuizSession.started_at.desc()).first()

    if existing and not existing.completed_at:
        eid = existing.id
        db.close()
        _prime_lb_cache(quiz_id)
        session.update({'session_id': eid, 'quiz_id': quiz_id, 'username': username, 'resuming': True})
        return redirect(url_for('quiz_page'))

    sess_id = str(uuid.uuid4())
    db.add(QuizSession(id=sess_id, quiz_id=quiz_id, username=username))
    db.commit()
    db.close()

    _prime_lb_cache(quiz_id)
    _lb_update(quiz_id, sess_id, username, 0, 0, 0)
    session.update({'session_id': sess_id, 'quiz_id': quiz_id, 'username': username, 'resuming': False})
    return redirect(url_for('quiz_page'))

@app.route('/quiz')
def quiz_page():
    if 'session_id' not in session:
        return redirect(url_for('index'))
    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == session['quiz_id']).first()
    qs = db.query(QuizSession).filter(QuizSession.id == session['session_id']).first()
    if not quiz or not qs:
        db.close()
        return redirect(url_for('index'))
    if qs.completed_at:
        db.close()
        return redirect(url_for('result'))
    active_q_id = quiz.active_question_id
    resuming = session.pop('resuming', False)
    db.close()
    return render_template('viva.html', quiz_id=quiz.id, username=session['username'],
                           active_question_id=active_q_id, resuming=resuming)


@app.route('/my_answers')
def my_answers():
    if 'session_id' not in session:
        return jsonify([])
    db = SessionLocal()
    answers = db.query(Answer).filter(Answer.session_id == session['session_id']).all()
    db.close()
    return jsonify([a.question_id for a in answers])

@app.route('/submit', methods=['POST'])
def submit():
    if 'session_id' not in session:
        return jsonify({'error': 'not logged in'}), 401

    data = request.get_json()
    q_id = int(data.get('question_id'))
    user_answer = data.get('answer', '')
    time_taken = float(data.get('time_taken', 0))
    sess_id = session['session_id']
    quiz_id = session['quiz_id']
    username = session['username']

    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz or q_id >= len(quiz.questions_data):
        db.close()
        return jsonify({'error': 'Invalid'}), 400

    existing_ans = db.query(Answer).filter(
        Answer.session_id == sess_id, Answer.question_id == q_id).first()
    if existing_ans:
        spd = existing_ans.speed_points
        ic = existing_ans.is_correct
        db.close()
        return jsonify({'ok': True, 'is_correct': ic, 'already': True, 'speed_points': spd})

    q = quiz.questions_data[q_id]
    db.close()

    is_correct, partial_score, any_wrong = check_answer_detailed(q, user_answer)
    spd = calc_speed_points(time_taken, q.get('time_limit')) if not any_wrong else 0

    with _lb_lock:
        entry = _lb_cache[quiz_id].get(sess_id, {'username': username, 'correct': 0, 'speed_points': 0, 'strikes': 0})
    _lb_update(quiz_id, sess_id, username,
               entry['correct'] + (1 if is_correct else 0),
               entry['speed_points'] + spd, entry['strikes'])

    answer_id, now = str(uuid.uuid4()), datetime.utcnow()

    def _write():
        db2 = SessionLocal()
        try:
            db2.add(Answer(id=answer_id, quiz_id=quiz_id, session_id=sess_id,
                           question_id=q_id, user_answer=user_answer,
                           is_correct=is_correct, partial_score=partial_score,
                           speed_points=spd, time_taken=time_taken, answered_at=now))
            qs = db2.query(QuizSession).filter(QuizSession.id == sess_id).first()
            if qs:
                if is_correct:
                    qs.correct_count += 1
                qs.speed_points += spd
            db2.commit()
        except Exception as e:
            db2.rollback()
            logger.error(f"Submit write: {e}")
        finally:
            db2.close()

    _write_queue.put(_write)
    return jsonify({'ok': True, 'is_correct': is_correct, 'speed_points': spd})

@app.route('/strike', methods=['POST'])
def log_strike():
    if 'session_id' not in session:
        return jsonify({'strikes': 0}), 401

    data = request.get_json(silent=True) or {}
    q_id = data.get('question_id')
    sess_id = session['session_id']
    quiz_id = session['quiz_id']
    username = session['username']

    with _lb_lock:
        entry = _lb_cache[quiz_id].get(sess_id, {'username': username, 'correct': 0, 'speed_points': 0, 'strikes': 0})
        new_strikes = entry['strikes'] + 1
        _lb_cache[quiz_id][sess_id] = {**entry, 'strikes': new_strikes}
        _lb_pending.add(quiz_id)

    answer_id, now = str(uuid.uuid4()), datetime.utcnow()

    def _write():
        db2 = SessionLocal()
        try:
            qs = db2.query(QuizSession).filter(QuizSession.id == sess_id).first()
            if qs:
                qs.strikes += 1
            if q_id is not None:
                exists = db2.query(Answer).filter(
                    Answer.session_id == sess_id, Answer.question_id == int(q_id)).first()
                if not exists:
                    db2.add(Answer(id=answer_id, quiz_id=quiz_id, session_id=sess_id,
                                   question_id=int(q_id), user_answer='[strike]',
                                   is_correct=False, partial_score=0, speed_points=0,
                                   time_taken=None, strike_on_question=True, answered_at=now))
            db2.commit()
        except Exception as e:
            db2.rollback()
            logger.error(f"Strike write: {e}")
        finally:
            db2.close()

    _write_queue.put(_write)
    return jsonify({'strikes': new_strikes, 'terminate': False})

@app.route('/result')
def result():
    if 'session_id' not in session:
        return redirect(url_for('index'))
    db = SessionLocal()
    qs = db.query(QuizSession).filter(QuizSession.id == session['session_id']).first()
    quiz = db.query(Quiz).filter(Quiz.id == session['quiz_id']).first()
    if qs and not qs.completed_at:
        qs.completed_at = datetime.utcnow()
        db.commit()
    correct = qs.correct_count if qs else 0
    total = len(quiz.questions_data) if quiz else 0
    speed = qs.speed_points if qs else 0
    db.close()
    session.clear()
    return (f"<!DOCTYPE html><html><body style='font-family:sans-serif;text-align:center;"
            f"margin-top:80px;background:#111;color:#fff'><h1>Quiz Complete!</h1>"
            f"<p>Correct: {correct}/{total}</p><p>Speed Points: {speed}</p>"
            f"<p style='color:#888'>Results recorded. You may close this tab.</p></body></html>")

# ─── ROUTES - ADMIN ────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin_id'] = str(uuid.uuid4())
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html', error='Incorrect password.')
    return render_template('admin_login.html', error=None)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    db = SessionLocal()
    quizzes = db.query(Quiz).order_by(Quiz.created_at.desc()).all()
    data = [{'id': q.id, 'title': q.title, 'pin': q.pin, 'state': q.state,
             'questions_data': q.questions_data, 'active_question_id': q.active_question_id}
            for q in quizzes]
    db.close()
    return render_template('admin.html', quizzes=data)

@app.route('/admin/quizzes')
def admin_quizzes():
    if not session.get('admin'):
        return '', 403
    db = SessionLocal()
    quizzes = db.query(Quiz).all()
    data = [{'id': q.id, 'state': q.state, 'active_question_id': q.active_question_id} for q in quizzes]
    db.close()
    return jsonify(data)

@app.route('/admin/quiz/new', methods=['POST'])
def create_quiz():
    if not session.get('admin'):
        return '', 403
    title = request.form.get('title', 'Untitled Quiz')
    try:
        questions = json.loads(request.form.get('questions_json', '[]'))
    except:
        questions = []
    pin, quiz_id = gen_pin(), str(uuid.uuid4())
    join_url = request.host_url.rstrip('/') + f'/?pin={pin}'
    qr_b64 = gen_qr(join_url)
    db = SessionLocal()
    db.add(Quiz(id=quiz_id, title=title, pin=pin,
                admin_id=session.get('admin_id', 'admin'),
                questions_data=questions, state='idle'))
    db.commit()
    db.close()
    return jsonify({'quiz_id': quiz_id, 'pin': pin, 'qr': qr_b64, 'join_url': join_url})

@app.route('/admin/quiz/<quiz_id>/question/open/<int:q_id>', methods=['POST'])
def open_question(quiz_id, q_id):
    if not session.get('admin'):
        return '', 403
    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if not quiz or q_id >= len(quiz.questions_data):
        db.close()
        return '', 400
    qdata = quiz.questions_data[q_id]
    quiz.active_question_id = q_id
    quiz.active_question_open_time = datetime.utcnow()
    quiz.state = 'running'
    db.commit()
    db.close()
    socketio.emit('question_opened', {'question_id': q_id, 'question': qdata}, room=quiz_id)
    return jsonify({'ok': True})

@app.route('/admin/quiz/<quiz_id>/question/close', methods=['POST'])
def close_question(quiz_id):
    if not session.get('admin'):
        return '', 403
    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    if quiz:
        correct_answer = None
        if quiz.active_question_id is not None and quiz.questions_data:
            q = quiz.questions_data[quiz.active_question_id]
            correct_answer = _get_correct_answer_str(q)
        quiz.active_question_id = None
        quiz.state = 'idle'
        db.commit()
        socketio.emit('question_closed', {'correct_answer': correct_answer}, room=quiz_id)
    db.close()
    return jsonify({'ok': True})

@app.route('/admin/quiz/<quiz_id>/stats')
def quiz_stats(quiz_id):
    if not session.get('admin'):
        return '', 403
    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    sessions = db.query(QuizSession).filter(QuizSession.quiz_id == quiz_id).all()
    questions = quiz.questions_data if quiz else []
    stats = []
    for sess in sessions:
        answers = db.query(Answer).filter(Answer.session_id == sess.id).all()
        ans_map = {a.question_id: a for a in answers}
        stats.append({
            'username': sess.username, 'correct_count': sess.correct_count,
            'speed_points': sess.speed_points, 'total_questions': len(questions),
            'strikes': sess.strikes, 'attended': sess.attended,
            'questions': [{
                'q_id': i, 'q_text': q.get('q', ''), 'q_type': q.get('type', 'mcq'),
                'user_answer': ans_map[i].user_answer if i in ans_map else None,
                'correct_answer': _get_correct_answer_str(q),
                'is_correct': ans_map[i].is_correct if i in ans_map else None,
                'speed_points': ans_map[i].speed_points if i in ans_map else 0,
                'time_taken': ans_map[i].time_taken if i in ans_map else None,
                'strike_on_question': ans_map[i].strike_on_question if i in ans_map else False,
            } for i, q in enumerate(questions)]
        })
    db.close()
    return jsonify(stats)

@app.route('/admin/quiz/<quiz_id>/export')
def quiz_export(quiz_id):
    if not session.get('admin'):
        return '', 403
    import io as _io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        USE_XLSX = True
    except ImportError:
        USE_XLSX = False

    db = SessionLocal()
    quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
    sessions = db.query(QuizSession).filter(QuizSession.quiz_id == quiz_id).all()
    questions = quiz.questions_data if quiz else []
    all_data = []
    for sess in sessions:
        answers = db.query(Answer).filter(Answer.session_id == sess.id).all()
        all_data.append((sess, {a.question_id: a for a in answers}))
    db.close()

    hdr = ["Username", "Total Correct", "Speed Points", "Strikes"]
    for i in range(len(questions)):
        hdr += [f"Q{i+1} Result", f"Q{i+1} Speed Pts"]

    def make_row(sess, ans_map):
        row = [sess.username, sess.correct_count, sess.speed_points, sess.strikes]
        for i in range(len(questions)):
            a = ans_map.get(i)
            if not a: row += ['N/A', 0]
            elif a.strike_on_question: row += ['Strike', 0]
            else: row += ['Correct' if a.is_correct else 'Wrong', a.speed_points]
        return row

    def q_summary():
        rows = []
        for i, q in enumerate(questions):
            correct = wrong = unanswered = 0
            for sess, ans_map in all_data:
                a = ans_map.get(i)
                if not a: unanswered += 1
                elif a.is_correct: correct += 1
                else: wrong += 1
            total = correct + wrong
            rows.append([i+1, q.get('q',''), q.get('type',''), correct, wrong, unanswered,
                         f"{round(correct/total*100,1)}%" if total else "0%"])
        return rows

    if USE_XLSX:
        wb = openpyxl.Workbook()
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="1E3A5F")

        ws1 = wb.active
        ws1.title = "Student Details"
        ws1.append(hdr)
        for cell in ws1[1]:
            cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
        for sess, ans_map in all_data:
            ws1.append(make_row(sess, ans_map))

        ws2 = wb.create_sheet("Question Summary")
        ws2.append(["Q#", "Question", "Type", "Correct", "Wrong", "Not Attempted", "Correct %"])
        for cell in ws2[1]:
            cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center')
        for row in q_summary():
            ws2.append(row)

        buf = _io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'quiz_{quiz_id[:8]}_stats.xlsx')
    else:
        import csv as _csv
        out = _io.StringIO()
        w = _csv.writer(out)
        w.writerow(hdr)
        for sess, ans_map in all_data:
            w.writerow(make_row(sess, ans_map))
        w.writerow([])
        w.writerow(["Q#", "Question", "Type", "Correct", "Wrong", "Not Attempted", "Correct %"])
        for row in q_summary():
            w.writerow(row)
        return Response(out.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment;filename=quiz_stats.csv'})

# ─── WEBSOCKET ─────────────────────────────────────────────────────────────────

@socketio.on('join_quiz')
def on_join_quiz(data):
    quiz_id = data.get('quiz_id')
    join_room(quiz_id)
    _prime_lb_cache(quiz_id)
    emit('joined', {'ok': True})

@socketio.on('get_leaderboard')
def on_get_leaderboard(data):
    quiz_id = data.get('quiz_id')
    _prime_lb_cache(quiz_id)
    emit('leaderboard_update', {'leaderboard': _get_sorted_lb(quiz_id)})

if __name__ == '__main__':
    logger.info("Quiz Platform starting on 0.0.0.0:5001")
    socketio.run(app, host='0.0.0.0', port=5001, debug=False, allow_unsafe_werkzeug=True)